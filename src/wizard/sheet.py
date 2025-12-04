from pathlib import Path
import string
from typing import Self, overload, Optional, Any

from pydantic import Field, field_validator, model_validator
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
import xlsxwriter
import pandas as pd
import ezodf
from ezodf.table import Table

from wizard.base import Serializable
from wizard.cell import Cell, DataType
from wizard.utils import timeit, get_logger, resolve_python_to_ezodf
from wizard.utils.google.client import Client, BackOffClient

logger = get_logger(__name__)


class Sheet(Serializable):
    """Represents a single sheet in a spreadsheet.

    The sheet contains two columns: one for original data and another for calculated data automatically converted by the spreadsheet software.
    """

    title: str | int
    sheet: pd.DataFrame = Field(default_factory=pd.DataFrame)
    parent: Optional["Book"] = Field(default=None, exclude=True, repr=False)

    @field_validator("sheet", mode="before")
    def sheet_validator(cls, sheet: pd.DataFrame) -> pd.DataFrame:
        # for i in range(sheet.shape[0]):
        #     for j in range(sheet.shape[1]):
        #         cell = sheet.iat[i, j]
        #         if not isinstance(cell, Cell):
        #             sheet.iat[i, j] = Cell(value=cell)
        #             cell = sheet.iat[i, j]
        #         cell.row = i + 1
        #         cell.column = j + 1
        return sheet

    @classmethod
    @timeit
    def from_excel(cls, title: str, dsheet: Worksheet, fsheet: Worksheet) -> Self:
        """Load a sheet from data and formula worksheets."""

        dataframe = pd.DataFrame()
        row1st = next(dsheet.iter_rows())
        # Find the last cell with a non-empty value in the first row
        try:
            ncell = list(filter(lambda cell: cell.value is not None, row1st))[-1]
        except IndexError:
            raise ValueError("The first row of sheet must not be empty")

        for colid, col in enumerate(dsheet.iter_cols(max_col=ncell.column), 1):
            dataframe[colid] = [
                Cell.from_excel(dcell, fsheet.cell(rowid, colid))
                for rowid, dcell in enumerate(col, 1)
            ]

        sheet = cls(title=title, sheet=dataframe)
        sheet.sheet.map(lambda cell: setattr(cell, "parent", sheet))
        return sheet

    @classmethod
    def from_ods(cls, sheet: Table) -> Self:
        """Load a sheet from an Table."""

        datas = [[Cell.from_ods(cell) for cell in row] for row in sheet.rows()]

        dataframe = pd.DataFrame(datas, columns=range(1, sheet.ncols() + 1))
        sheet = cls(title=sheet.name, sheet=dataframe)
        sheet.sheet.map(lambda cell: setattr(cell, "parent", sheet))
        return sheet

    @classmethod
    def from_gsheet(
        cls,
        title: str | int,
        dsheet: list[list[Any]],
        fsheet: Optional[list[list[Any]]] = None,
    ) -> Self:
        if fsheet is None:
            datas = list(
                map(lambda row: [Cell.from_gsheet(col) for col in row], dsheet)
            )
        else:
            datas = []
            # fmt: off
            for rowid, row in enumerate(dsheet, 1):
                if len(row) != len(fsheet[rowid - 1]):
                    Warning(f"Mismatch in length: Raw data at row {row} differs from rendered data. {row} -> {fsheet[rowid - 1]}. Skipping...")
                    continue

                datas.append([
                    Cell.from_gsheet(value, fsheet[rowid - 1][colid - 1])
                    for colid, value in enumerate(row, 1)
                ])
            # fmt: on

        dataframe = pd.DataFrame(
            datas, columns=[i for i in range(1, len(datas[0]) + 1)]
        )
        sheet = cls(title=title, sheet=dataframe)
        return sheet

    @overload
    def __getitem__(self, index: tuple[int, int]) -> Cell:
        """Retrieves a single cell based on its 1-based row and column index."""
        ...

    @overload
    def __getitem__(self, row: int) -> pd.Series:
        """Retrieves a single row of cells."""
        ...

    @overload
    def __getitem__(self, index: tuple[int, slice]) -> Cell:
        """Retrieves a single row of cells and specific columns based on their 1-based row and column indices."""
        ...

    @overload
    def __getitem__(self, index: tuple[slice, int]) -> Cell:
        """Retrieves a single column of cells and specific rows based on their 1-based row and column indices."""
        ...

    @overload
    def __getitem__(self, index: tuple[slice, slice]) -> pd.DataFrame:
        """Retrieves a subset of the sheet based on the specified 1-based row and column slices."""
        ...

    @overload
    def __getitem__(self, index: str) -> Cell | pd.Series:
        """Retrieves either a single cell using a string index (e.g., 'A1')
        or an entire column of cells based on the specified index."""
        ...

    @overload
    def __getitem__(self, index: slice) -> pd.DataFrame:
        """Retrieves multiple rows of cells."""
        ...

    def __getitem__(
        self, index: tuple[int, int] | str | int | slice
    ) -> Cell | pd.Series | pd.DataFrame:
        def refine_index(index: int | slice) -> int | slice:
            if isinstance(index, slice):
                start = None if index.start is None else refine_index(index.start)
                stop = None if index.stop is None else refine_index(index.stop)
                return slice(start, stop, index.step)
            elif isinstance(index, int):
                if index == 0:
                    raise IndexError("Index must be one-based.")
                return index if index < 0 else index - 1

        if isinstance(index, tuple):
            row, col = index
            # a subset of the sheet
            return self.sheet.iloc[refine_index(row), refine_index(col)]

        if isinstance(index, str):
            # a single column
            if len(index) == 1:
                return self.sheet.iloc[:, string.ascii_uppercase.index(index)]

            # a single cell
            col, row = index[:1], index[1:]
            return self.sheet.iat[
                refine_index(int(row)), string.ascii_uppercase.index(col)
            ]

        # a single row or multiple rows
        if isinstance(index, (int, slice)):
            return self.sheet.iloc[refine_index(index)]

        raise IndexError("Invalid index")

    @overload
    def __setitem__(self, index: tuple[int, int], cell: Cell):
        """Sets a single cell based on its row and column(1-based)."""
        ...

    @overload
    def __setitem__(self, index: str, cell: Cell):
        """Sets a single cell based on str index, such as A1."""
        ...

    def __setitem__(self, index: tuple[int, int] | str, cell: Cell):
        row, col = None, None
        if isinstance(index, tuple):
            row, col = index
        elif isinstance(index, str):
            col, row = string.ascii_uppercase.index(index[:1]) + 1, int(index[1:])

        # check if the row and column are within the bounds
        if row and col and row <= len(self.nrows) and col <= len(self.ncols):
            self.sheet.iloc[row - 1, col - 1] = cell
        else:
            raise ValueError("Invalid index")

    @property
    def nrows(self) -> int:
        """Returns the number of rows in the sheet."""
        return self.sheet.shape[0]

    @property
    def ncols(self) -> int:
        """Returns the number of columns in the sheet."""
        return self.sheet.shape[1]

    def __len__(self) -> int:
        return self.nrows

    def to_book(self, path: Optional[Path] = None) -> "Book":
        """Convert the sheet to a spreadsheet."""
        return Book.from_sheet(self, path)

    def split(self, size: int) -> list[Self]:
        """Splits the sheet into multiple sheets."""
        if size >= len(self):
            return [self]

        retvals = []
        for idx, i in enumerate(range(0, len(self), size)):
            df = self.sheet.iloc[i : i + size]
            retvals.append(
                self.__class__(
                    title=f"{self.title}_{idx}", sheet=df.reset_index(drop=True)
                )
            )

        return retvals


class Book(Serializable):
    """Represents a spreadsheet containing multiple sheets.

    Args:
        uid: The spreadsheet's unique identifier. For a book loaded from a local Excel file, the file path serves as the identifier. For books loaded from the Google Docs online service, the spreadsheet ID is used as the identifier.
    """

    uid: Path | str
    title: Optional[str] = None
    sheets: dict[str | int, Sheet] = Field(default_factory=dict)
    active_sheet: Sheet = Field(default=None)

    @model_validator(mode="before")
    def active_sheet_present(cls, d: dict):
        """Verify that `active_sheet` is included in the `sheets` collection."""
        active_sheet_present = False
        for sheet in d["sheets"].values():
            if id(sheet) == id(d["active_sheet"]):
                active_sheet_present = True

        if not active_sheet_present:
            raise ValueError("active_sheet must be in sheets")
        return d

    def model_post_init(self, __context: Any) -> None:
        """Set the parent of each sheet to the instance."""
        self._set_sheets_parent()
        if self.title is None:
            if isinstance(self.uid, Path):
                self.title = self.uid.stem
            else:
                self.title = self.uid

    @classmethod
    def from_excel(cls, path: Path) -> Self:
        """Loads a spreadsheet from a local Excel file."""

        logger.info(f"Loading spreadsheet from {path}")

        dbook = load_workbook(str(path), data_only=True)
        fbook = load_workbook(str(path), data_only=False)

        sheets = {}
        active_sheet = None
        for dsheet, fsheet in zip(dbook, fbook):
            obj = Sheet.from_excel(dsheet.title, dsheet, fsheet)
            sheets[dsheet.title] = obj

            if dsheet == dbook.active:
                active_sheet = obj

        dbook.close()
        fbook.close()

        return cls(uid=path, title=path.stem, sheets=sheets, active_sheet=active_sheet)

    @classmethod
    def from_calc(cls, path: Path) -> Self:
        """Loads a spreadsheet from a local ODS file."""
        sheets, active_sheet = {}, None
        book = ezodf.opendoc(path)
        for dsheet in book.sheets:
            sheet = Sheet.from_ods(dsheet)
            sheets[dsheet.name] = sheet

            if active_sheet is None:
                active_sheet = sheet

        return cls(uid=path, title=path.stem, sheets=sheets, active_sheet=active_sheet)

    @classmethod
    def from_gsheet(
        cls,
        id: str,
        credential_path: Optional[str] = None,
        backoff: bool = True,
        optimized_for_rate_limit: bool = False,
        load_formula: bool = True,
    ):

        client = (
            BackOffClient.from_credential_path(credential_path)
            if backoff
            else Client.from_credential_path(credential_path)
        )

        spreadsheet = client.open(id)
        dsheets = (
            spreadsheet.get_data_by_single_request(formula=False)
            if optimized_for_rate_limit
            else spreadsheet.get_data_by_multiple_requests(formula=False)
        )
        if load_formula:
            fsheets = (
                spreadsheet.get_data_by_single_request(formula=True)
                if optimized_for_rate_limit
                else spreadsheet.get_data_by_multiple_requests(formula=True)
            )
        else:
            fsheets = {key: None for key in dsheets.keys()}

        sheets = {
            title: Sheet.from_gsheet(title, dsheet, fsheet)
            for title, (dsheet, fsheet) in zip(
                dsheets.keys(), zip(dsheets.values(), fsheets.values())
            )
        }

        return cls(
            uid=spreadsheet.uid,
            title=None,
            sheets=sheets,
            active_sheet=next(iter(sheets.values())),
        )

    def to_gsheet(
        self,
        folder_id: Optional[str] = None,
        credential_path: Optional[str] = None,
        backoff: bool = True,
        range_format: Optional[dict[str, dict[str, Any]]] = None,
    ) -> str:
        from wizard.utils.google.client import Client, BackOffClient

        client = (
            BackOffClient.from_credential_path(credential_path)
            if backoff
            else Client.from_credential_path(credential_path)
        )

        spreadsheet = client.create(self.title, folder_id=folder_id)
        sheets = {
            title: sheet.sheet.map(lambda cell: cell.content).to_numpy().tolist()
            for title, sheet in self.sheets.items()
        }
        spreadsheet.update(sheets, range_format=range_format)
        return spreadsheet.uid

    def to_xlsx(self, path: Optional[Path] = None) -> Path:
        """Saves the spreadsheet to the file specified by 'path', or to its original location if no path is provided."""
        if path is None:
            path = self.uid

        logger.info(f"Saving spreadsheet to {path}")
        options = {
            "strings_to_formulas": False,
            "strings_to_urls": False,
            "default_date_format": "yyyy-mm-dd hh:mm:ss",
        }
        workbook = xlsxwriter.Workbook(path, options=options)

        for sheet in self.sheets.values():
            worksheet = workbook.add_worksheet(sheet.title)
            for i, row in sheet.sheet.iterrows():
                for j, cell in enumerate(row):
                    if cell.is_formula():
                        worksheet.write_formula(i, j, cell.formula)
                    else:
                        if (
                            cell.datatype == DataType.STRING
                            and cell.content.startswith("=")
                        ):
                            worksheet.write_string(i, j, cell.content)
                        elif cell.datatype == DataType.DATETIME:
                            worksheet.write_datetime(i, j, cell.content)
                        else:
                            worksheet.write(i, j, cell.content)

        workbook.close()
        return path

    def to_ods(self, path: Optional[Path] = None) -> Path:
        """Saves the spreadsheet to the file specified by 'path', or to its original location if no path is provided."""
        if path is None:
            path = self.uid

        book = ezodf.newdoc(doctype="ods", filename=path)
        book.backup = False

        for title, sheet in self.sheets.items():
            table = Table(title, size=(sheet.nrows, sheet.ncols))
            for i, row in sheet.sheet.iterrows():
                for j, cell in enumerate(row):
                    # skip empty cells
                    # FIXME(gpl): what about cells with formulas?
                    if cell.content is not None:
                        value, value_type = resolve_python_to_ezodf(cell.content)
                        table[i, j].set_value(value, value_type=value_type)
            book.sheets += table
            book.save()

        return path

    def add_sheet(self, sheet: Sheet):
        """Adds a sheet to the spreadsheet."""
        self.sheets[sheet.title] = sheet
        if self.active_sheet is None:
            self.active_sheet = sheet

        sheet.parent = self

    def split(self, size: int) -> list[Self]:
        """Splits the spreadsheet into multiple single sheet spreadsheets."""
        retvals = []
        parent, stem, suffix = self.uid.parent, self.uid.stem, self.uid.suffix
        for sheet in self.sheets.values():
            retvals.extend(
                map(
                    lambda x: self.__class__.from_sheet(
                        x, parent / (stem + x.title + suffix)
                    ),
                    sheet.split(size),
                )
            )
        return retvals

    @classmethod
    def from_sheet(cls, sheet: Sheet, uid: Optional[Path | str] = None) -> Self:
        """Creates a spreadsheet from a single sheet."""
        uid = uid or Path(sheet.title)
        return cls(
            uid=uid, title=sheet.title, sheets={sheet.title: sheet}, active_sheet=sheet
        )

    def _set_sheets_parent(self):
        for sheet in self.sheets.values():
            sheet.parent = self
